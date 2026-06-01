"""
onboarding — org onboarding state machine + bulk-user Excel parser.

Org-RBAC overhaul (Lane C). Drives Organization.onboarding_state through a
strict linear sequence:

    created
      -> ai_configured
      -> storage_connected
      -> domains_created
      -> users_added
      -> completed

`assert_step_allowed` enforces ordering (422 if a transition is attempted out of
order). `advance_state` persists the new state (idempotent: advancing to the
current state or backward is a no-op so re-running a step never regresses).

`parse_users_xlsx` reads a bulk-upload .xlsx into normalized user rows; it is
tolerant of header-name variations and of domains given as a comma/semicolon/
pipe separated string or as separate columns.
"""
from __future__ import annotations

import io
import re
from datetime import datetime, timezone

import structlog
from fastapi import HTTPException
from sqlalchemy.ext.asyncio import AsyncSession

logger = structlog.get_logger("onboarding")

# ── State machine ──────────────────────────────────────────────────────────

# Ordered list of onboarding states. Index = progress depth.
ONBOARDING_STATES: list[str] = [
    "created",
    "ai_configured",
    "storage_connected",
    "domains_created",
    "users_added",
    "completed",
]

# step name (what a transition produces) -> the state it requires to already be in
# i.e. to reach STATE you must currently be at the immediately-preceding state.
_STATE_INDEX = {state: i for i, state in enumerate(ONBOARDING_STATES)}


def _current_index(org) -> int:
    state = getattr(org, "onboarding_state", None) or "created"
    return _STATE_INDEX.get(state, 0)


def assert_step_allowed(org, target_step: str) -> None:
    """Raise 422 unless `target_step` is a valid next (or already-reached) state
    for `org`. Re-applying the current state is allowed (idempotent). Skipping
    ahead is rejected.

    Rule: target index must be <= current index + 1 (no skipping), and must be a
    known state.
    """
    if target_step not in _STATE_INDEX:
        raise HTTPException(
            status_code=422, detail=f"Unknown onboarding step '{target_step}'"
        )
    cur = _current_index(org)
    target = _STATE_INDEX[target_step]
    if target > cur + 1:
        expected = ONBOARDING_STATES[cur + 1] if cur + 1 < len(ONBOARDING_STATES) else None
        raise HTTPException(
            status_code=422,
            detail=(
                f"Onboarding step out of order: org is at "
                f"'{ONBOARDING_STATES[cur]}', cannot jump to '{target_step}'. "
                f"Next allowed step is '{expected}'."
            ),
        )


async def advance_state(org, step: str, db: AsyncSession) -> str:
    """Advance org.onboarding_state to `step` (forward-only) and persist.

    Idempotent: if the org is already at or beyond `step`, the state is left
    unchanged. When advancing to 'completed', stamps onboarding_completed_at.
    Returns the resulting state.
    """
    assert_step_allowed(org, step)
    target = _STATE_INDEX[step]
    if target > _current_index(org):
        org.onboarding_state = step
        if step == "completed" and hasattr(org, "onboarding_completed_at"):
            org.onboarding_completed_at = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(org)
        logger.info("onboarding_state_advanced", org_id=org.id, state=step)
    return org.onboarding_state


def build_checklist(org) -> dict[str, bool]:
    """Return per-step completion booleans for the GET /state endpoint."""
    cur = _current_index(org)
    return {
        "ai_configured": cur >= _STATE_INDEX["ai_configured"],
        "storage_connected": cur >= _STATE_INDEX["storage_connected"],
        "domains_created": cur >= _STATE_INDEX["domains_created"],
        "users_added": cur >= _STATE_INDEX["users_added"],
        "completed": cur >= _STATE_INDEX["completed"],
    }


# ── Bulk-user Excel parser ───────────────────────────────────────────────────

_EMAIL_HEADERS = {"email", "e-mail", "email_address", "emailaddress", "mail", "user_email"}
_ROLE_HEADERS = {"role", "user_role", "access_level", "type"}
_DOMAIN_HEADERS = {
    "domains",
    "domain",
    "allowed_domains",
    "department",
    "departments",
    "domain_tags",
    "domain_tag",
}
_NAME_HEADERS = {"name", "full_name", "fullname", "display_name", "username", "user"}

_SPLIT_RE = re.compile(r"[,;|/]")
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _norm_header(value: object) -> str:
    return re.sub(r"[\s\-]+", "_", str(value or "").strip().lower())


def _split_domains(value: object) -> list[str]:
    if value is None:
        return []
    parts = _SPLIT_RE.split(str(value))
    return [p.strip() for p in parts if p and p.strip()]


def parse_users_xlsx(file_bytes: bytes) -> list[dict]:
    """Parse a bulk-user .xlsx into a list of normalized dicts:

        {"email": str, "role": str | None, "domains": list[str], "name": str | None}

    Tolerant of header variations (Email / E-mail / mail; Role / Access Level;
    Domains / Department / Allowed Domains). Domains may be a single delimited
    cell (comma/semicolon/pipe/slash) or spread across repeated domain columns.
    The first row is treated as the header. Rows without a valid email are
    skipped. Raises 422 if no email column is found or no valid rows parse.
    """
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500, detail=f"openpyxl unavailable: {exc}"
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=422, detail=f"Could not read .xlsx file: {exc}"
        ) from exc

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Spreadsheet is empty")

    headers = [_norm_header(h) for h in header_row]

    email_idx: int | None = None
    role_idx: int | None = None
    name_idx: int | None = None
    domain_idxs: list[int] = []
    for i, h in enumerate(headers):
        if email_idx is None and h in _EMAIL_HEADERS:
            email_idx = i
        elif role_idx is None and h in _ROLE_HEADERS:
            role_idx = i
        elif name_idx is None and h in _NAME_HEADERS:
            name_idx = i
        elif h in _DOMAIN_HEADERS:
            domain_idxs.append(i)

    # Fallback: if no labelled email column, sniff for a column whose first data
    # value looks like an email.
    if email_idx is None:
        raise HTTPException(
            status_code=422,
            detail="No email column found in spreadsheet (expected a header like 'email').",
        )

    out: list[dict] = []
    for raw in rows_iter:
        if raw is None:
            continue
        cells = list(raw)

        def _at(idx: int | None):
            if idx is None or idx >= len(cells):
                return None
            return cells[idx]

        email = _at(email_idx)
        email = str(email).strip().lower() if email is not None else ""
        if not email or not _EMAIL_RE.match(email):
            continue

        role = _at(role_idx)
        role = str(role).strip().lower() if role else None

        name = _at(name_idx)
        name = str(name).strip() if name else None

        domains: list[str] = []
        for di in domain_idxs:
            domains.extend(_split_domains(_at(di)))
        # de-dup, preserve order
        seen: set[str] = set()
        domains = [d for d in domains if not (d in seen or seen.add(d))]

        out.append({"email": email, "role": role, "domains": domains, "name": name})

    wb.close()

    if not out:
        raise HTTPException(
            status_code=422, detail="No valid user rows found in spreadsheet."
        )
    return out
