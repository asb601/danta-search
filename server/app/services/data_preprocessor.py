"""
Data preprocessing pipeline — runs BEFORE DuckDB sampling and Parquet conversion.

Handles every supported file type:
    .csv  .txt  .tsv  .tab  .dat  .psv  .pipe  .xlsx  .xls  .xlsm  .xlsb

Resource model (zero-disk for CSV, half-disk for Excel)
────────────────────────────────────────────────────────
File type   Disk used during preprocessing       RAM (peak)
─────────   ──────────────────────────────       ──────────
CSV / text  ZERO  — probe bytes via HTTP range   one chunk  (~200 MB)
            read, stream full file from Azure,   regardless of
            write clean blocks directly back      file size
            to Azure. No local file ever written.

Excel       ONE temp file (the workbook download). Openpyxl/xlrd need a
            seekable file. Excel jobs are single-flight per worker process and
            fail before download if tmp free space is below the safety margin.
            Rows are streamed to Azure block blob — no second local clean file.

Memory model
────────────
Large files (> 50 MB) are streamed 50 000 rows at a time. Small files
(<= 50 MB) are kept in RAM to allow exact-duplicate removal.

Cleaning stages (per-chunk for large files, once for small)
────────────────────────────────────────────────────────────
  1.  Encoding detection   — HTTP range read of first 64 KB only
  2.  Delimiter detection  — first 8 KB of probe bytes (clevercsv → extended
       Sniffer → frequency+consistency analysis; any single-char delimiter)
  3.  Header detection     — first 15 rows of probe bytes
  4.  Schema discovery     — first 1 000 rows of probe bytes → per-column type
  5.  Per-chunk streaming:
        a. String cleaning  — BOM / control-chars / invisible-unicode / whitespace
        b. Null normalise   — "", "NULL", "N/A", "nan", ... -> ""
        c. Garbage-row drop — subtotals, separator lines, fully-empty rows
        d. Per-column conv  — bool / date / numeric normalisation
  6.  Duplicate removal    — ONLY for small files; skipped for large
  7.  Output              — written as Azure Block Blob (staged 4 MB blocks)

Output: clean UTF-8 CSV.
        - CSV / text  →  same blob path, overwritten in place (Azure atomic block commit).
    - Excel       →  sibling .cleaned-<file_id>.csv blob next to the workbook.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import io
import os
import re
import shutil
import tempfile
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, date as _date_type
from pathlib import Path
from typing import Callable, Iterator

import pandas as pd
from azure.storage.blob import BlobClient, BlobServiceClient

from app.core.logger import ingest_logger
from app.services.preprocessor.cleaning_rules import CleaningProfile, get_cleaning_profile


# ── Tunable constants ─────────────────────────────────────────────────────────

CHUNK_ROWS              = 50_000        # rows per streaming chunk (lower = less RAM per thread)
SMALL_FILE_THRESHOLD_MB = 50            # files under this get full-load + dedup
HEADER_SCAN_ROWS        = 15            # max rows to scan for the real header
TYPE_DETECT_SAMPLE_ROWS = 1_000         # rows used for column-type detection
PROBE_BYTES             = 256 * 1024    # bytes range-read from Azure for probing
BLOCK_SIZE              = 4 * 1024 * 1024  # 4 MB per Azure block blob block
MIN_EXCEL_TMP_FREE_BYTES = 2 * 1024 * 1024 * 1024  # keep 2 GB free after download
EXCEL_TMP_FREE_MULTIPLIER = 2.0  # require free tmp space >= max(2 GB, file_size * 2)


# ── Supported file-type groups ────────────────────────────────────────────────

EXCEL_EXTS = frozenset({".xlsx", ".xls", ".xlsm", ".xlsb"})
TEXT_EXTS  = frozenset({
    # Standard delimited-text formats
    ".csv", ".tsv", ".tab",
    # Generic text — any delimiter detected automatically
    ".txt",
    # ERP / data-warehouse exports often use these extensions
    ".dat",   # SAP, Oracle EBS generic data export
    ".psv",   # pipe-separated values
    ".pipe",  # pipe-separated (alternate convention)
    ".dsv",   # delimiter-separated values
})
ALL_EXTS   = EXCEL_EXTS | TEXT_EXTS


# ── Null-like string patterns (compared after .strip().lower()) ───────────────

_NULLSTR: frozenset[str] = frozenset({
    "", "null", "none", "na", "n/a", "nan", "nil", "tbd", "n.a.", "n.a",
    "-", "--", "---", ".", "..", "?", "#", "#n/a", "#na", "#null!",
    "not available", "not applicable", "not provided", "not assigned",
    "missing", "unknown", "no data", "no value", "nd", "n.d.",
    "void", "blank", "empty",
})


# ── Compiled regex patterns ───────────────────────────────────────────────────
# Type-detection regex (currency, thousands, percent, Excel date serials, hint
# tokens) lives in app.services.preprocessor.type_detection — owned by the
# detectors that use them. Patterns kept here are used by row / cell cleaning
# (garbage-row removal, control-char stripping) and the header-row scanner.

_GARBAGE_ROW_RE = re.compile(
    # NOTE: pandas' string accessor (`.str.match`/`.str.contains`) on Arrow-
    # backed string columns routes the pattern through PyArrow's RE2 compute
    # kernel.  RE2 does NOT accept Python-style `\uXXXX` escapes inside
    # character classes — it raises `Invalid regular expression: invalid
    # escape sequence: \u`.  Build the non-ASCII pieces from literal Unicode
    # codepoints (concatenated outside a raw string) so RE2 sees the raw
    # character, not the escape sequence.
    r"^\s*(total|grand\s+total|subtotal|sub\s+total|sum|page\s+total|"
    r"running\s+total|end\s+of\s+report|average|avg|mean|balance\s+forward|"
    r"carried\s+forward|min|max|"
    # German (SAP, other ERP exports)
    r"summe|gesamtsumme|gesamt|zwischensumme|durchschnitt|"
    # French
    r"total\s+g[e" + "\u00e9" + r"]n[e" + "\u00e9" + r"]ral|total\s+partiel|moyenne|"
    # Spanish / Portuguese
    r"total\s+general|suma\s+total|promedio|m[e" + "\u00e9" + r"]dia)\b",
    re.IGNORECASE,
)
_SEP_ROW_RE    = re.compile(r"^[-=*_~\s|+]+$")
# Same RE2-safety rule as _GARBAGE_ROW_RE: build with literal codepoints, not
# `\xNN` / `\uXXXX` escapes, so the pattern works whether pandas evaluates it
# natively or via PyArrow.
_CTRL_RE       = re.compile(
    "["
    "\x00-\x08"
    "\x0b\x0c"
    "\x0e-\x1f"
    "\x7f"
    "]"
)
_INVISIBLE_RE  = re.compile(
    "["
    "\u200b\u200c\u200d\u200e\u200f\ufeff\u00ad"
    "\u180e\u2060\u2061\u2062\u2063\u2064\u3000"
    "]"
)
# Pre-compiled patterns used by vectorized _clean_str_series
_NEWLINE_RE     = re.compile(r"[\r\n\t]")
_MULTI_SPACE_RE = re.compile(r" {2,}")

# ── Concurrency guard ─────────────────────────────────────────────────────────
# Preprocessing is CPU + I/O heavy. On an 8 GB VM, allow at most 2 concurrent
# preprocessing jobs so pandas string ops don't saturate all cores / exhaust RAM.
_PREPROCESS_SEMAPHORE: asyncio.Semaphore = asyncio.Semaphore(2)
_EXCEL_PREPROCESS_SEMAPHORE: asyncio.Semaphore = asyncio.Semaphore(1)


def _get_preprocess_semaphore() -> asyncio.Semaphore:
    """Return the process-wide semaphore used to cap preprocessing RAM."""
    return _PREPROCESS_SEMAPHORE


def _get_excel_preprocess_semaphore() -> asyncio.Semaphore:
    """Return the process-wide semaphore used to cap local Excel temp disk."""
    return _EXCEL_PREPROCESS_SEMAPHORE


# ── Result dataclass ──────────────────────────────────────────────────────────

@dataclass
class PreprocessResult:
    clean_blob_path: str
    original_rows:   int
    clean_rows:      int
    rows_dropped:    int
    cols_renamed:    dict = field(default_factory=dict)
    warnings:        list = field(default_factory=list)
    encoding:        str  = "utf-8"
    file_type:       str  = "csv"
    used_streaming:  bool = False
    already_clean:   bool = False  # True = probe showed no changes needed, full scan skipped
    # ── Cleaning audit — populated by the pluggable cleaning-rule registry ──
    # Total rows removed during garbage/empty/separator detection.
    quarantine_count:  int  = 0
    # First MAX_QUARANTINE_SAMPLE dropped rows with their reason — stored in
    # FileAnalytics.quarantine_sample for ops audit without reprocessing the file.
    quarantine_sample: list = field(default_factory=list)
    cleaning_audit: dict = field(default_factory=dict)


# ══════════════════════════════════════════════════════════════════════════════
# Azure streaming helpers
# ══════════════════════════════════════════════════════════════════════════════

class _AzureRawStream(io.RawIOBase):
    """
    Wraps an Azure StorageStreamDownloader as a readable io.RawIOBase so
    pandas can read it directly via pd.read_csv(stream).

    No data is written to disk.  The Azure SDK fetches data over HTTP in
    chunks; this class stitches those chunks together into the read()
    interface that pandas expects.
    """

    def __init__(self, downloader) -> None:
        self._chunks   = downloader.chunks()
        self._leftover = b""

    def readable(self) -> bool:
        return True

    def readinto(self, b: bytearray) -> int:
        want = len(b)
        # Keep pulling Azure chunks until we have enough bytes or stream ends
        while len(self._leftover) < want:
            try:
                self._leftover += next(self._chunks)
            except StopIteration:
                break
        n = min(want, len(self._leftover))
        if n == 0:
            return 0
        b[:n] = self._leftover[:n]
        self._leftover = self._leftover[n:]
        return n


class _BlockBlobWriter:
    """
    Accumulates bytes in an in-memory buffer and uploads to Azure Blob Storage
    using the Block Blob API (stage_block + commit_block_list).

    Why Block Blob?
        • We can upload arbitrarily large outputs without holding the full
          content in memory or on disk.
        • Each block is at most BLOCK_SIZE bytes, so peak extra RAM is
          bounded by one block (~4 MB).
        • commit() finalises the blob atomically.

    Usage:
        writer = _BlockBlobWriter(blob_client)
        writer.write(b"some bytes")
        ...
        writer.commit()   # must be called exactly once at the end
    """

    def __init__(self, bc: BlobClient) -> None:
        self._bc     = bc
        self._buf    = bytearray()
        self._blocks: list[str] = []
        self._idx    = 0

    def write(self, data: bytes) -> None:
        self._buf.extend(data)
        while len(self._buf) >= BLOCK_SIZE:
            self._flush_block(bytes(self._buf[:BLOCK_SIZE]))
            del self._buf[:BLOCK_SIZE]

    def _flush_block(self, data: bytes) -> None:
        block_id = base64.b64encode(f"{self._idx:08d}".encode()).decode()
        self._bc.stage_block(block_id=block_id, data=data)
        self._blocks.append(block_id)
        self._idx += 1

    def commit(self) -> None:
        """Flush remaining buffer and commit all staged blocks."""
        if self._buf:
            self._flush_block(bytes(self._buf))
            self._buf = bytearray()
        if self._blocks:
            self._bc.commit_block_list(self._blocks)
        else:
            # Nothing was written — produce an empty blob with just the header
            self._bc.upload_blob(b"", overwrite=True)


# ══════════════════════════════════════════════════════════════════════════════
# Public entry point
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ProbeResult:
    """Lightweight result of the 256KB probe — no full download."""
    safe_for_raw_sample: bool   # True = DuckDB can sample the raw file reliably
    encoding:           str     # detected encoding
    header_row_idx:     int     # 0 = header is first row; >0 = junk rows precede it
    reason:             str     # why it's not safe (empty when safe)


async def probe_raw_csv(
    blob_path:         str,
    file_name:         str,
    connection_string: str,
    container_name:    str,
) -> ProbeResult:
    """
    Range-read the first 256 KB of a CSV/text blob and determine whether
    DuckDB can reliably sample it without preprocessing first.

    Unsafe conditions (require preprocessing before sampling):
      1. Non-UTF-8 encoding  — DuckDB reads garbled/wrong strings
      2. Leading junk rows   — DuckDB uses the junk as column names,
                               making the AI description completely wrong

    Everything else (dirty null strings like "N/A", whitespace, control
    chars inside values) does not materially affect the AI description
    and can be tolerated for the sample.
    """
    ext = Path(file_name).suffix.lower()
    if ext not in TEXT_EXTS:
        # Excel files are never safe to sample raw
        return ProbeResult(safe_for_raw_sample=False, encoding="",
                           header_row_idx=0, reason="excel")

    def _run() -> ProbeResult:
        svc = BlobServiceClient.from_connection_string(connection_string)
        bc  = svc.get_blob_client(container=container_name, blob=blob_path)
        raw = _probe_blob(bc, PROBE_BYTES)

        encoding = _detect_encoding_from_bytes(raw)
        # UTF-8-sig is fine for DuckDB (it handles the BOM)
        clean_encoding = encoding in ("utf-8", "utf-8-sig", "ascii")
        if not clean_encoding:
            return ProbeResult(
                safe_for_raw_sample=False,
                encoding=encoding,
                header_row_idx=0,
                reason=f"non-utf8 encoding: {encoding}",
            )

        probe_text = raw.decode(encoding, errors="replace")
        delimiter  = _detect_delimiter_from_str(probe_text, ext)
        head_df = pd.read_csv(
            io.StringIO(probe_text), sep=delimiter, header=None, dtype=str,
            keep_default_na=False, nrows=HEADER_SCAN_ROWS, on_bad_lines="skip",
        )
        head_df = head_df.apply(_clean_str_series).apply(_nullify_series)
        header_row_idx = _find_header_row(head_df)

        if header_row_idx > 0:
            return ProbeResult(
                safe_for_raw_sample=False,
                encoding=encoding,
                header_row_idx=header_row_idx,
                reason=f"{header_row_idx} leading junk row(s) before real header",
            )

        return ProbeResult(
            safe_for_raw_sample=True,
            encoding=encoding,
            header_row_idx=0,
            reason="",
        )

    return await asyncio.to_thread(_run)


async def preprocess_file(
    blob_path:         str,
    file_name:         str,
    file_id:           str,
    connection_string: str,
    container_name:    str,
    cleaning_config:   dict | None = None,
) -> PreprocessResult:
    """
    Clean a raw file and write clean CSV output back to Azure Blob Storage.

    cleaning_config: optional dict from ContainerConfig.cleaning_config.  When
        supplied, per-container null patterns and garbage row patterns are merged
        with the built-in registry — no code change or redeploy needed.
        Schema: {"extra_null_patterns": [...], "extra_garbage_patterns": [...]}

    Disk usage:
        CSV / text : ZERO disk.  Probe bytes via HTTP range-read (256 KB),
                     full file streamed directly from Azure, clean output
                     written as Azure Block Blob — no local file at any point.
        Excel      : ONE temp file (the .xlsx download; openpyxl needs
                     seekable I/O). Clean output still goes directly to Azure
                     as a Block Blob — no second local copy.

    RAM usage:
        Bounded by CHUNK_ROWS (100 000 rows ~ 200 MB) for any file size.
        For small files (<= 50 MB) the full dataframe is kept for dedup.

    Output blob:
        CSV / text  →  same path as input (overwritten in place atomically).
        Excel       →  sibling .csv blob (same folder, .xlsx → .csv).
    """
    t0 = time.perf_counter()
    warns: list[str] = []
    ext       = Path(file_name).suffix.lower()
    file_type = "excel" if ext in EXCEL_EXTS else "csv"

    ingest_logger.info("preprocess", status="started", blob_path=blob_path,
                       file_name=file_name, file_type=file_type)

    # ── Probe size without downloading the file ───────────────────────────────
    # Create BlobServiceClient once; reuse for src, dst, and properties checks.
    svc_client    = await asyncio.to_thread(
        BlobServiceClient.from_connection_string, connection_string
    )
    src_bc    = svc_client.get_blob_client(container=container_name, blob=blob_path)
    props     = await asyncio.to_thread(lambda: src_bc.get_blob_properties())
    file_size = props["size"]
    size_mb   = file_size / (1024 * 1024)
    is_large  = size_mb > SMALL_FILE_THRESHOLD_MB

    ingest_logger.info("preprocess", status="probed",
                       size_mb=round(size_mb, 1), streaming=is_large)

    # ── Build CleaningProfile from built-in rules + per-container extras ─────────────
    cfg             = cleaning_config or {}
    cleaning_profile = get_cleaning_profile(
        extra_null_patterns=cfg.get("extra_null_patterns", []),
        extra_garbage_re_patterns=cfg.get("extra_garbage_patterns", []),
    )
    # ── Output blob path: NEW INDUSTRY DESIGN ─────────────────────────────────
    # CSV / text → overwrite the SAME blob path in place.
    #              Azure block blob commit_block_list is atomic — readers continue
    #              to see the old content until the final commit, then the blob
    #              atomically swaps to the clean content.  Zero extra storage.
    # Excel      → write a sibling .csv next to the .xlsx (Excel cannot be
    #              overwritten with CSV; different format).  blob_path will
    #              be updated by the caller to point to the new .csv.
    if ext in EXCEL_EXTS:
        # report.xlsx → report.cleaned-<file_id>.csv (same folder)
        # Never write to report.csv: that may be a real user-uploaded file.
        clean_blob_path = blob_path[: -len(ext)] + f".cleaned-{file_id[:8]}.csv"
    else:
        # billing.csv → billing.csv  (overwritten in place, atomic)
        clean_blob_path = blob_path
    dst_bc          = svc_client.get_blob_client(container=container_name, blob=clean_blob_path)
    block_writer    = _BlockBlobWriter(dst_bc)

    # ── Semaphore: cap concurrent preprocessing to avoid OOM under heavy load ──
    _sem = _get_preprocess_semaphore()
    async with _sem:
        if ext in EXCEL_EXTS:
            # Excel still needs one local temp file because openpyxl/xlrd need
            # seekable input. Cap Excel concurrency separately and verify tmp
            # space before download so archived workbooks cannot fill the VM disk.
            _excel_sem = _get_excel_preprocess_semaphore()
            async with _excel_sem:
                with tempfile.TemporaryDirectory() as tmpdir:
                    _ensure_excel_tmp_capacity(tmpdir, file_size)
                    raw_path = os.path.join(tmpdir, f"raw{ext}")
                    await asyncio.to_thread(
                        _download_blob_to_file,
                        src_bc, raw_path,
                    )
                    result = await asyncio.to_thread(
                        _process_excel_to_blob, raw_path, block_writer, ext, is_large, warns,
                        cleaning_profile,
                    )
                    result["temp_disk_bytes"] = file_size
            # tmpdir (and the only temp file) is deleted here; block_writer still in RAM
        else:
            # CSV/text: fully streaming, zero disk
            result = await asyncio.to_thread(
                _process_text_stream,
                src_bc, block_writer, ext, file_size, is_large, warns,
                cleaning_profile,
            )

    ingest_logger.info("preprocess", status="cleaned",
                       original_rows=result["original_rows"],
                       clean_rows=result["clean_rows"],
                       rows_dropped=result["original_rows"] - result["clean_rows"],
                       quarantine_count=result.get("quarantine_count", 0),
                       cleaning_audit=result.get("cleaning_audit", {}),
                       streaming=is_large,
                       already_clean=result.get("already_clean", False))

    ingest_logger.info("preprocess", status="done",
                       clean_blob_path=clean_blob_path,
                       duration_ms=round((time.perf_counter() - t0) * 1000, 1))

    already_clean = result.get("already_clean", False)
    return PreprocessResult(
        clean_blob_path=clean_blob_path,
        original_rows=result["original_rows"],
        clean_rows=result["clean_rows"],
        rows_dropped=result["original_rows"] - result["clean_rows"],
        cols_renamed=result["cols_renamed"],
        warnings=warns,
        encoding=result.get("encoding", "utf-8"),
        file_type=file_type,
        used_streaming=is_large,
        already_clean=already_clean,
        quarantine_count=result.get("quarantine_count", 0),
        quarantine_sample=result.get("quarantine_sample", []),
        cleaning_audit={
            **(result.get("cleaning_audit") or {}),
            "file_type": file_type,
            "used_streaming": is_large,
            "temp_disk_bytes": result.get("temp_disk_bytes", 0),
            "clean_blob_path": clean_blob_path,
        },
    )


# ══════════════════════════════════════════════════════════════════════════════
# Azure client helpers
# ══════════════════════════════════════════════════════════════════════════════

def _get_blob_client(conn_str: str, container: str, blob_path: str,
                     _svc: "BlobServiceClient | None" = None) -> BlobClient:
    svc = _svc or BlobServiceClient.from_connection_string(conn_str)
    return svc.get_blob_client(container=container, blob=blob_path)


def _probe_blob(src_bc: BlobClient, length: int = PROBE_BYTES) -> bytes:
    """Range-read the first `length` bytes without downloading the entire blob."""
    actual = min(length, src_bc.get_blob_properties()["size"])
    return src_bc.download_blob(offset=0, length=actual).readall()


def _download_blob_to_file(bc: BlobClient, dest: str) -> None:
    """Full download to a local file (used only for Excel)."""
    with open(dest, "wb") as fh:
        bc.download_blob().readinto(fh)


def _ensure_excel_tmp_capacity(tmpdir: str, file_size: int) -> None:
    """Fail before downloading Excel if tmp disk cannot safely hold the file."""
    usage = shutil.disk_usage(tmpdir)
    required_free = max(
        MIN_EXCEL_TMP_FREE_BYTES,
        int(file_size * EXCEL_TMP_FREE_MULTIPLIER),
    )
    if usage.free < required_free:
        raise RuntimeError(
            "Insufficient temporary disk for Excel preprocessing: "
            f"free={usage.free} bytes, required={required_free} bytes, "
            f"file_size={file_size} bytes"
        )


# ══════════════════════════════════════════════════════════════════════════════
# CSV / text — fully streaming, zero disk
# ══════════════════════════════════════════════════════════════════════════════

def _process_text_stream(
    src_bc:       BlobClient,
    block_writer: _BlockBlobWriter,
    ext:          str,
    file_size:    int,
    is_large:     bool,
    warns:        list[str],
    profile:      CleaningProfile | None = None,
) -> dict:
    """
    CSV/text processing: Azure -> RAM chunks -> Azure block blob.
    No local file is written at any point.

    Pass 1 (probe): HTTP range-read of first PROBE_BYTES (256 KB) →
        detect encoding, delimiter, header row, column types.
    Pass 2 (full): full Azure stream → pd.read_csv(chunksize) →
        clean each chunk → stage as block blob blocks.
    """
    # ── Pass 1: probe ─────────────────────────────────────────────────────────
    probe      = _probe_blob(src_bc, PROBE_BYTES)
    encoding   = _detect_encoding_from_bytes(probe)
    probe_text = probe.decode(encoding, errors="replace")
    delimiter  = _detect_delimiter_from_str(probe_text, ext)

    head_df = pd.read_csv(
        io.StringIO(probe_text), sep=delimiter, header=None, dtype=str,
        keep_default_na=False, nrows=HEADER_SCAN_ROWS, on_bad_lines="skip",
    )
    head_df = head_df.apply(_clean_str_series).apply(_nullify_series)
    header_row_idx = _find_header_row(head_df)

    raw_headers = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(head_df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)
    skip_rows = header_row_idx + 1

    if header_row_idx > 0:
        warns.append(
            f"Header row found at row {header_row_idx} ({header_row_idx} leading rows skipped)"
        )

    sample_df = pd.read_csv(
        io.StringIO(probe_text), sep=delimiter, header=None, dtype=str, names=headers,
        keep_default_na=False, skiprows=skip_rows,
        nrows=TYPE_DETECT_SAMPLE_ROWS, on_bad_lines="skip",
    )
    sample_df  = sample_df.apply(_clean_str_series).apply(_nullify_series)
    converters = _build_converters(sample_df, headers, warns)

    # ── Fast path: skip full rewrite for large files that are already clean ────
    # For files >50 MB we check 5 probe conditions from Pass 1 data only.
    # If ALL met, the file is already a clean UTF-8 comma-delimited CSV with no
    # dirty null values → skip the 3-6 GB network round-trip entirely.
    #
    # Conditions:
    #   1. UTF-8 / ASCII encoding  (no re-encoding needed)
    #   2. Header at row 0         (no junk rows to skip)
    #   3. Delimiter is ','        (output is always comma)
    #   4. Column names unchanged  (no Unnamed: N, no dedup renames)
    #   5. Sample has no null-like strings  (e.g. 'N/A', 'null', 'none')
    if is_large:
        cell_strs      = [str(v) if v is not None else "" for v in head_df.iloc[header_row_idx]]
        headers_clean  = (cell_strs == raw_headers) and not cols_renamed
        encoding_clean = encoding in ("utf-8", "utf-8-sig", "ascii")
        delimiter_clean = delimiter == ","

        if encoding_clean and header_row_idx == 0 and delimiter_clean and headers_clean:
            # Re-read raw sample WITHOUT _nullify_series, compare to nullified version
            raw_sample = pd.read_csv(
                io.StringIO(probe_text), sep=delimiter, header=None, dtype=str,
                names=headers, keep_default_na=False, skiprows=1,
                nrows=TYPE_DETECT_SAMPLE_ROWS, on_bad_lines="skip",
            ).apply(_clean_str_series)
            sample_nullified = raw_sample.copy().apply(_nullify_series)
            no_null_changes  = raw_sample.fillna("").equals(sample_nullified.fillna(""))

            if no_null_changes:
                warns.append(
                    "fast_path: file is already clean UTF-8 CSV — "
                    "full rewrite skipped (probe passed all 5 checks)"
                )
                return {
                    "original_rows": 0,
                    "clean_rows":    0,
                    "cols_renamed":  cols_renamed,
                    "encoding":      encoding,
                    "already_clean": True,
                    "cleaning_audit": {
                        "header_row_idx": header_row_idx,
                        "delimiter": delimiter,
                        "dedup_skipped": False,
                        "rewrite_skipped": True,
                    },
                }

    # ── Pass 2: full streaming read ────────────────────────────────────────────
    _active_profile = profile if profile is not None else get_cleaning_profile()

    # Count malformed CSV rows (wrong column count) via on_bad_lines callable
    # (pandas >= 1.3) so we have an audit trail without a second file pass.
    _malformed_count: list[int] = [0]
    _malformed_sample: list[dict] = []

    def _handle_bad_line(bad_line: list) -> None:  # type: ignore[return]
        _malformed_count[0] += 1
        if len(_malformed_sample) < 5:
            _malformed_sample.append({
                "reason": "malformed_csv_row",
                "row": {"_raw": ",".join(str(x) for x in bad_line[:20])},
            })
        return None  # skip the row

    downloader = src_bc.download_blob()
    raw_stream = io.BufferedReader(_AzureRawStream(downloader), buffer_size=8 * 1024 * 1024)

    try:
        reader = pd.read_csv(
            raw_stream, sep=delimiter, header=None, dtype=str, names=headers,
            encoding=encoding, encoding_errors="replace",
            keep_default_na=False, skiprows=skip_rows,
            chunksize=CHUNK_ROWS, on_bad_lines=_handle_bad_line,
        )
    except TypeError:
        # pandas < 1.3 — on_bad_lines callable not supported; fall back silently
        raw_stream.seek(0)
        reader = pd.read_csv(
            raw_stream, sep=delimiter, header=None, dtype=str, names=headers,
            encoding=encoding, encoding_errors="replace",
            keep_default_na=False, skiprows=skip_rows,
            chunksize=CHUNK_ROWS, on_bad_lines="skip",
        )

    # Write CSV header row as the first block
    header_bytes = (",".join(headers) + "\n").encode("utf-8")
    block_writer.write(header_bytes)

    original_rows     = 0
    clean_rows        = 0
    total_quarantine  = 0
    quarantine_sample: list[dict] = []
    small_chunks: list[pd.DataFrame] = []

    for chunk in reader:
        original_rows += len(chunk)
        chunk, q_rows  = _clean_chunk(chunk, converters, _active_profile)
        total_quarantine += len(q_rows)
        if len(quarantine_sample) < 20:
            quarantine_sample.extend(q_rows[:20 - len(quarantine_sample)])

        if is_large:
            buf = io.StringIO()
            chunk.to_csv(buf, index=False, header=False)
            block_writer.write(buf.getvalue().encode("utf-8"))
            clean_rows += len(chunk)
        else:
            small_chunks.append(chunk)

    if not is_large and small_chunks:
        full   = pd.concat(small_chunks, ignore_index=True)
        before = len(full)
        full   = full.drop_duplicates()
        n_dup  = before - len(full)
        if n_dup:
            warns.append(f"Dropped {n_dup} exact-duplicate row(s)")
        full = full.fillna("")
        buf  = io.StringIO()
        full.to_csv(buf, index=False, header=False)
        block_writer.write(buf.getvalue().encode("utf-8"))
        clean_rows = len(full)
    elif is_large:
        warns.append("Deduplication skipped for large file to keep memory bounded")

    block_writer.commit()

    # Merge malformed CSV rows into the quarantine audit
    if _malformed_count[0]:
        warns.append(f"Skipped {_malformed_count[0]} malformed CSV row(s) (wrong column count)")
        total_quarantine += _malformed_count[0]
        if len(quarantine_sample) < 20:
            quarantine_sample.extend(_malformed_sample[:20 - len(quarantine_sample)])

    return {
        "original_rows":     original_rows,
        "clean_rows":        clean_rows,
        "cols_renamed":      cols_renamed,
        "encoding":          encoding,
        "already_clean":     False,
        "quarantine_count":  total_quarantine,
        "quarantine_sample": quarantine_sample,
        "cleaning_audit": {
            "header_row_idx": header_row_idx,
            "delimiter": delimiter,
            "dedup_skipped": is_large,
            "rewrite_skipped": False,
        },
    }


def _clean_chunk(
    chunk: pd.DataFrame,
    converters: dict,
    profile: CleaningProfile,
) -> tuple[pd.DataFrame, list[dict]]:
    """Apply all per-row / per-cell cleaning to a single chunk.

    Returns (clean_chunk, quarantine_sample).
    quarantine_sample: list of dicts [{reason, row}] for dropped rows.
    """
    # Cell-level: string normalisation then null-pattern replacement via registry
    chunk = chunk.apply(_clean_str_series).apply(profile.nullify_series)
    chunk = chunk.dropna(how="all")
    # Row-level: garbage / empty / separator detection via pluggable registry
    chunk, quarantine = profile.clean_rows(chunk)
    for col, fn in converters.items():
        if col in chunk.columns:
            try:
                chunk[col] = chunk[col].apply(fn)
            except Exception:
                pass
    return chunk.fillna(""), quarantine


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — one temp file download, direct Azure block blob output
# ══════════════════════════════════════════════════════════════════════════════

def _process_excel_to_blob(
    raw_path:     str,
    block_writer: _BlockBlobWriter,
    ext:          str,
    is_large:     bool,
    warns:        list[str],
    profile:      CleaningProfile | None = None,
) -> dict:
    if ext in (".xlsx", ".xlsm"):
        return _process_xlsx_to_blob(raw_path, block_writer, is_large, warns, profile)
    return _process_xls_to_blob(raw_path, block_writer, warns, profile)


def _process_xlsx_to_blob(
    raw_path:     str,
    block_writer: _BlockBlobWriter,
    is_large:     bool,
    warns:        list[str],
    profile:      CleaningProfile | None = None,
) -> dict:
    """
    Stream .xlsx via openpyxl read_only=True, write directly to Azure block blob.
    Peak disk: only the .xlsx download (no second clean file).
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string  # noqa: PLC0415

    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    # Prefer the sheet with the most populated cells — the active sheet is often
    # a cover page or instructions tab, not the actual data.
    _data_sheets   = [s for s in wb.worksheets if (s.max_row or 0) > 0]
    ws             = max(_data_sheets, key=lambda s: (s.max_row or 0) * (s.max_column or 0)) \
                     if _data_sheets else wb.active
    best_sheet_title = ws.title

    hidden_col_idx: set[int] = set()
    try:
        for col_letter, col_dim in ws.column_dimensions.items():
            if col_dim.hidden:
                try:
                    hidden_col_idx.add(column_index_from_string(col_letter))
                except Exception:
                    pass
    except Exception:
        pass

    def _iter_visible_rows(ws_obj) -> Iterator[list[str]]:
        for row in ws_obj.iter_rows():
            row_num = row[0].row if row else None
            if row_num is None:
                continue
            try:
                rd = ws_obj.row_dimensions.get(row_num)
                if rd and rd.hidden:
                    continue
            except Exception:
                pass
            yield [
                str(cell.value) if cell.value is not None else ""
                for j, cell in enumerate(row, start=1)
                if j not in hidden_col_idx
            ]

    # Collect header-scan rows
    head_buf: list[list[str]] = []
    row_iter = _iter_visible_rows(ws)
    for raw_row in row_iter:
        head_buf.append(raw_row)
        if len(head_buf) >= HEADER_SCAN_ROWS:
            break
    wb.close()

    if not head_buf:
        warns.append("Excel file appears empty")
        block_writer.write(b"")
        block_writer.commit()
        return {
            "original_rows": 0,
            "clean_rows": 0,
            "cols_renamed": {},
            "encoding": "binary",
            "cleaning_audit": {"empty_excel": True},
        }

    head_df        = pd.DataFrame(head_buf).astype(str)
    head_df        = head_df.apply(_clean_str_series).apply(_nullify_series)
    header_row_idx = _find_header_row(head_df)
    raw_headers    = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(head_df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)
    n_cols        = len(headers)
    data_leftover = head_buf[header_row_idx + 1:]

    # Re-open for the full streaming pass (use the same sheet identified above)
    wb2 = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    ws2 = wb2[best_sheet_title] if best_sheet_title in wb2.sheetnames else wb2.active

    def _iter_data_rows() -> Iterator[list[str]]:
        for r in data_leftover:
            yield r[:n_cols] + [""] * max(0, n_cols - len(r))
        skipped   = 0
        skip_limit = header_row_idx + 1
        for row in ws2.iter_rows():
            row_num = row[0].row if row else None
            if row_num is None:
                continue
            try:
                rd = ws2.row_dimensions.get(row_num)
                if rd and rd.hidden:
                    continue
            except Exception:
                pass
            if skipped < skip_limit:
                skipped += 1
                continue
            vals = [
                str(cell.value) if cell.value is not None else ""
                for j, cell in enumerate(row, start=1)
                if j not in hidden_col_idx
            ]
            yield vals[:n_cols] + [""] * max(0, n_cols - len(vals))

    # Sample for type detection
    sample_rows: list[list[str]] = []
    data_gen = _iter_data_rows()
    for row in data_gen:
        sample_rows.append(row)
        if len(sample_rows) >= TYPE_DETECT_SAMPLE_ROWS:
            break

    sample_df  = pd.DataFrame(sample_rows, columns=headers).astype(str)
    sample_df  = sample_df.apply(_clean_str_series).apply(_nullify_series)
    converters = _build_converters(sample_df, headers, warns)

    _active_profile = profile if profile is not None else get_cleaning_profile()

    # Write header to block blob
    block_writer.write((",".join(headers) + "\n").encode("utf-8"))

    original_rows    = len(sample_rows)
    clean_rows       = 0
    total_quarantine = 0
    quarantine_sample: list[dict] = []
    small_chunks: list[pd.DataFrame] = []

    def _make_chunk(rows: list[list[str]]) -> tuple[pd.DataFrame, list[dict]]:
        df = pd.DataFrame(rows, columns=headers).astype(str)
        return _clean_chunk(df, converters, _active_profile)

    sample_clean, q_rows = _make_chunk(sample_rows)
    total_quarantine += len(q_rows)
    if len(quarantine_sample) < 20:
        quarantine_sample.extend(q_rows[:20 - len(quarantine_sample)])
    if is_large:
        buf = io.StringIO()
        sample_clean.to_csv(buf, index=False, header=False)
        block_writer.write(buf.getvalue().encode("utf-8"))
        clean_rows += len(sample_clean)
    else:
        small_chunks.append(sample_clean)

    batch: list[list[str]] = []
    for row in data_gen:
        original_rows += 1
        batch.append(row)
        if len(batch) >= CHUNK_ROWS:
            chunk, q_rows = _make_chunk(batch)
            total_quarantine += len(q_rows)
            if len(quarantine_sample) < 20:
                quarantine_sample.extend(q_rows[:20 - len(quarantine_sample)])
            if is_large:
                buf = io.StringIO()
                chunk.to_csv(buf, index=False, header=False)
                block_writer.write(buf.getvalue().encode("utf-8"))
                clean_rows += len(chunk)
            else:
                small_chunks.append(chunk)
            batch = []

    if batch:
        chunk, q_rows = _make_chunk(batch)
        total_quarantine += len(q_rows)
        if len(quarantine_sample) < 20:
            quarantine_sample.extend(q_rows[:20 - len(quarantine_sample)])
        if is_large:
            buf = io.StringIO()
            chunk.to_csv(buf, index=False, header=False)
            block_writer.write(buf.getvalue().encode("utf-8"))
            clean_rows += len(chunk)
        else:
            small_chunks.append(chunk)

    if not is_large and small_chunks:
        full   = pd.concat(small_chunks, ignore_index=True)
        before = len(full)
        full   = full.drop_duplicates()
        n_dup  = before - len(full)
        if n_dup:
            warns.append(f"Dropped {n_dup} exact-duplicate row(s)")
        full = full.fillna("")
        buf  = io.StringIO()
        full.to_csv(buf, index=False, header=False)
        block_writer.write(buf.getvalue().encode("utf-8"))
        clean_rows = len(full)
    elif is_large:
        warns.append("Deduplication skipped for large Excel file to keep memory bounded")

    block_writer.commit()
    wb2.close()

    return {
        "original_rows":     original_rows,
        "clean_rows":        clean_rows,
        "cols_renamed":      cols_renamed,
        "encoding":          "binary",
        "quarantine_count":  total_quarantine,
        "quarantine_sample": quarantine_sample,
        "cleaning_audit": {
            "sheet": best_sheet_title,
            "header_row_idx": header_row_idx,
            "dedup_skipped": is_large,
            "rewrite_skipped": False,
        },
    }


def _process_xls_to_blob(
    raw_path:     str,
    block_writer: _BlockBlobWriter,
    warns:        list[str],
    profile:      CleaningProfile | None = None,
) -> dict:
    """
    .xls legacy format via xlrd + pandas.
    xlrd doesn't support streaming but .xls is capped at 65 535 rows.
    Clean output written directly to block blob — no second local file.
    """
    try:
        df = pd.read_excel(raw_path, header=None, dtype=object, engine="xlrd")
    except ImportError:
        warns.append("xlrd not installed; .xls support unavailable")
        block_writer.commit()
        return {
            "original_rows": 0,
            "clean_rows": 0,
            "cols_renamed": {},
            "encoding": "binary",
            "cleaning_audit": {"xls_read_failed": "xlrd_missing"},
        }
    except Exception as ex:
        warns.append(f".xls read failed: {ex}")
        block_writer.commit()
        return {
            "original_rows": 0,
            "clean_rows": 0,
            "cols_renamed": {},
            "encoding": "binary",
            "cleaning_audit": {"xls_read_failed": str(ex)[:300]},
        }

    original_rows  = len(df)
    _active_profile = profile if profile is not None else get_cleaning_profile()
    df = df.astype(str).apply(_clean_str_series).apply(_active_profile.nullify_series)

    header_row_idx = _find_header_row(df.iloc[:HEADER_SCAN_ROWS])
    raw_headers    = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)

    df = df.iloc[header_row_idx + 1:].copy()
    ncols = len(df.columns)
    df.columns = headers[:ncols] + [f"col_{i}" for i in range(ncols - len(headers))]
    converters = _build_converters(df.iloc[:TYPE_DETECT_SAMPLE_ROWS], headers, warns)
    df, q_rows = _clean_chunk(df, converters, _active_profile)

    before = len(df)
    df = df.drop_duplicates()
    n_dup = before - len(df)
    if n_dup:
        warns.append(f"Dropped {n_dup} exact-duplicate row(s)")

    buf = io.StringIO()
    df.fillna("").to_csv(buf, index=False)
    block_writer.write(buf.getvalue().encode("utf-8"))
    block_writer.commit()

    return {
        "original_rows":     original_rows,
        "clean_rows":        len(df),
        "cols_renamed":      cols_renamed,
        "encoding":          "binary",
        "quarantine_count":  len(q_rows),
        "quarantine_sample": q_rows,
        "cleaning_audit": {
            "header_row_idx": header_row_idx,
            "dedup_skipped": False,
            "rewrite_skipped": False,
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# Encoding + delimiter detection
# ══════════════════════════════════════════════════════════════════════════════

def _detect_encoding_from_bytes(raw: bytes) -> str:
    """Detect encoding from a bytes object (e.g. probe data already in memory)."""
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if raw[:4] in (b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff"):
        return "utf-32"
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return "utf-16"
    try:
        from charset_normalizer import from_bytes as _cnb  # noqa: PLC0415
        best = _cnb(raw).best()
        if best:
            return str(best.encoding)
    except ImportError:
        pass
    for enc in ("utf-8", "cp1252", "iso-8859-1", "latin-1"):
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            pass
    return "utf-8"


def _detect_encoding(path: str) -> str:
    """Detect encoding from a local file path (used by the file-path test helpers)."""
    with open(path, "rb") as fh:
        return _detect_encoding_from_bytes(fh.read(65536))


def _is_consistent_delimiter(text: str, delim: str, threshold: float = 0.80) -> bool:
    """
    Return True if *delim* produces a consistent column count across lines.
    >80% of non-empty lines must split into the same number of fields.
    """
    lines = [ln for ln in text.splitlines() if ln.strip()][:60]
    if len(lines) < 2:
        return True  # single-line files can't be validated; accept
    counts = [ln.count(delim) for ln in lines]
    mode_count = max(set(counts), key=counts.count)
    if mode_count == 0:
        return False
    agree = sum(1 for c in counts if c == mode_count)
    return agree / len(counts) >= threshold


def _frequency_delimiter(text: str) -> str | None:
    """
    Rank candidate delimiters by (frequency × consistency) and return the best one.
    Score favours characters that appear often *and* consistently across all lines.
    Space is given a lower weight because it appears in almost every cell value.
    """
    # Ordered roughly by real-world prevalence in data files
    _CANDIDATES = [",", "\t", "|", ";", ":", "^", "~", "!", " "]
    lines = [ln for ln in text.splitlines() if ln.strip()][:60]
    if len(lines) < 2:
        return None

    best_delim: str | None = None
    best_score = 0.0

    for delim in _CANDIDATES:
        counts = [ln.count(delim) for ln in lines]
        mode_count = max(set(counts), key=counts.count)
        if mode_count == 0:
            continue
        agree = sum(1 for c in counts if c == mode_count) / len(counts)
        # Space penalised — virtually every line contains spaces that aren't delimiters
        weight = 0.4 if delim == " " else 1.0
        score = mode_count * agree * weight
        if score > best_score and agree >= 0.75 and mode_count >= 1:
            best_score = score
            best_delim = delim

    # Space requires extra evidence: at least 3 consistent fields per line
    if best_delim == " ":
        counts = [ln.count(" ") for ln in lines]
        mode_count = max(set(counts), key=counts.count)
        if mode_count < 2:
            return None

    return best_delim


def _detect_delimiter_from_str(text: str, ext_hint: str = "") -> str:
    """
    Detect the field delimiter from decoded text using a four-level cascade:

    1. clevercsv.Sniffer  — statistical analysis; handles any single-char delimiter
       including space, colon, caret, tilde, etc.  (installed via pyproject.toml)
    2. csv.Sniffer        — extended candidate set beyond the stdlib default
    3. Frequency + consistency analysis — manual scoring across the first 60 lines
    4. Extension hint     — last-resort fallback for well-known extension conventions

    The probe sample used is the first 8 KB (sufficient for sniffing without reading
    the whole file into memory).
    """
    sample = text[:8192]

    # ── 1. clevercsv (preferred) ───────────────────────────────────────────────
    try:
        import clevercsv  # noqa: PLC0415
        dialect = clevercsv.Sniffer().sniff(sample, verbose=False)
        if dialect is not None:
            return dialect.delimiter
    except Exception:
        pass

    # ── 2. csv.Sniffer with extended candidates ────────────────────────────────
    _SNIFFER_CANDIDATES = ",\t|;:^~! "
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=_SNIFFER_CANDIDATES)
        delim = dialect.delimiter
        if _is_consistent_delimiter(sample, delim):
            return delim
    except csv.Error:
        pass

    # ── 3. Frequency + consistency scoring ────────────────────────────────────
    best = _frequency_delimiter(sample)
    if best is not None:
        return best

    # ── 4. Extension hint as absolute last resort ──────────────────────────────
    if ext_hint.endswith((".tsv", ".tab")):
        return "\t"
    if ext_hint.endswith((".psv", ".pipe")):
        return "|"
    return ","


def _detect_delimiter(path: str, encoding: str) -> str:
    """Detect CSV delimiter from a local file path (used by the file-path test helpers)."""
    try:
        with open(path, encoding=encoding, errors="replace") as fh:
            return _detect_delimiter_from_str(fh.read(8192), path)
    except OSError:
        return ","


# ══════════════════════════════════════════════════════════════════════════════
# File-path based processing (kept for tests + local fallback)
# ══════════════════════════════════════════════════════════════════════════════

def _process_text(
    raw_path:   str,
    clean_path: str,
    is_large:   bool,
    warns:      list[str],
) -> dict:
    """
    File-path based CSV processing (used in unit tests and local fallback).
    For production Azure ingestion, _process_text_stream is used instead.
    """
    encoding  = _detect_encoding(raw_path)
    delimiter = _detect_delimiter(raw_path, encoding)

    head_df = pd.read_csv(
        raw_path, sep=delimiter, header=None, dtype=str,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, nrows=HEADER_SCAN_ROWS,
        on_bad_lines="skip",
    )
    head_df = head_df.apply(_clean_str_series).apply(_nullify_series)
    header_row_idx = _find_header_row(head_df)

    raw_headers = [
        _flatten_col_name(v) or f"col_{i}"
        for i, v in enumerate(head_df.iloc[header_row_idx])
    ]
    headers, cols_renamed = _dedup_column_names(raw_headers)
    skip_rows = header_row_idx + 1

    if header_row_idx > 0:
        warns.append(f"Header row found at row {header_row_idx} ({header_row_idx} leading rows skipped)")

    sample_df = pd.read_csv(
        raw_path, sep=delimiter, header=None, dtype=str, names=headers,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, skiprows=skip_rows,
        nrows=TYPE_DETECT_SAMPLE_ROWS, on_bad_lines="skip",
    )
    sample_df = sample_df.apply(_clean_str_series).apply(_nullify_series)
    converters = _build_converters(sample_df, headers, warns)

    original_rows = 0
    clean_rows    = 0
    small_chunks: list[pd.DataFrame] = []

    reader = pd.read_csv(
        raw_path, sep=delimiter, header=None, dtype=str, names=headers,
        encoding=encoding, encoding_errors="replace",
        keep_default_na=False, skiprows=skip_rows,
        chunksize=CHUNK_ROWS, on_bad_lines="skip",
    )

    with open(clean_path, "w", encoding="utf-8", newline="") as out_fh:
        writer = csv.writer(out_fh, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)

        for chunk in reader:
            original_rows += len(chunk)
            chunk = _clean_chunk(chunk, converters)

            if is_large:
                for row in chunk.itertuples(index=False, name=None):
                    writer.writerow(row)
                clean_rows += len(chunk)
            else:
                small_chunks.append(chunk)

        if not is_large and small_chunks:
            full = pd.concat(small_chunks, ignore_index=True)
            before = len(full)
            full   = full.drop_duplicates()
            n_dup  = before - len(full)
            if n_dup:
                warns.append(f"Dropped {n_dup} exact-duplicate row(s)")
            full = full.fillna("")
            for row in full.itertuples(index=False, name=None):
                writer.writerow(row)
            clean_rows = len(full)

    return {
        "original_rows": original_rows,
        "clean_rows":    clean_rows,
        "cols_renamed":  cols_renamed,
        "encoding":      encoding,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Cell-level string cleaning
# ══════════════════════════════════════════════════════════════════════════════

def _clean_str_series(s: pd.Series) -> pd.Series:
    """Vectorized column cleaning — one C-level call per pattern, NOT a Python loop.

    Replaces the old `s.apply(_clean_str)` which fired a Python function call
    for every single cell (30M+ calls on a SAP file with 300 cols × 100K rows).
    Vectorized .str operations call into pandas/regex C extensions once per column.
    """
    # Only process object/string columns — pass through already-typed numerics/dates
    if s.dtype.kind not in ("O", "U"):
        return s
    # Remove invisible Unicode and control characters
    s = s.str.replace(_INVISIBLE_RE, "", regex=True)
    s = s.str.replace(_CTRL_RE, "", regex=True)
    # Collapse line endings and tabs to a single space
    s = s.str.replace(_NEWLINE_RE, " ", regex=True)
    # Collapse repeated spaces and strip leading/trailing whitespace
    s = s.str.replace(_MULTI_SPACE_RE, " ", regex=True)
    s = s.str.strip()
    # Normalise empty strings to NaN so downstream .fillna("") works correctly
    s = s.replace("", None)
    return s


def _nullify_series(s: pd.Series) -> pd.Series:
    """Vectorized null normalization — one isin() call for all null-like strings.

    Replaces the old `s.apply(_nullify)` per-cell Python loop.
    """
    if s.dtype.kind not in ("O", "U"):
        return s
    lower = s.astype(str).str.strip().str.lower()
    null_mask = lower.isin(_NULLSTR) | s.isna()
    return s.where(~null_mask, other=None)


# ══════════════════════════════════════════════════════════════════════════════
# Header detection (reads only first HEADER_SCAN_ROWS)
# ══════════════════════════════════════════════════════════════════════════════

def _is_numeric_str(v: str) -> bool:
    try:
        float(str(v).strip().replace(",", "").replace("_", ""))
        return True
    except (ValueError, TypeError):
        return False


def _find_header_row(df: pd.DataFrame) -> int:
    max_scan  = min(HEADER_SCAN_ROWS, len(df))
    best_row, best_score = 0, -1.0

    for i in range(max_scan):
        row      = df.iloc[i]
        non_null = [v for v in row if v is not None]
        if not non_null:
            continue
        str_cnt   = sum(1 for v in non_null if isinstance(v, str) and not _is_numeric_str(v))
        num_cnt   = sum(1 for v in non_null if isinstance(v, str) and _is_numeric_str(v))
        coverage  = len(non_null) / max(len(row), 1)
        str_ratio = str_cnt / max(len(non_null), 1)
        num_ratio = num_cnt / max(len(non_null), 1)
        avg_len   = sum(len(str(v)) for v in non_null) / max(len(non_null), 1)
        len_pen   = max(0.0, (avg_len - 60) / 150)
        score = coverage * str_ratio - num_ratio * 0.5 - len_pen

        if score > best_score + 0.05:
            best_score = score
            best_row   = i
        if i <= 3 and score > 0.70:
            break

    return best_row


# ══════════════════════════════════════════════════════════════════════════════
# Structural helpers
# ══════════════════════════════════════════════════════════════════════════════

def _flatten_col_name(v: object) -> str:
    if isinstance(v, tuple):
        parts = [str(p).strip() for p in v
                 if p is not None and not str(p).lower().startswith("unnamed")]
        return "_".join(parts) if parts else ""
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s*Unnamed:\s*\d+(_level_\d+)?", "", s, flags=re.IGNORECASE).strip()
    return s


def _dedup_column_names(names: list[str]) -> tuple[list[str], dict]:
    renamed: dict[str, str] = {}
    seen:    dict[str, int] = {}
    result:  list[str]      = []
    for col in names:
        clean = col.strip() or "col"
        if clean in seen:
            seen[clean] += 1
            new = f"{clean}_{seen[clean]}"
            renamed[col] = new
            result.append(new)
        else:
            seen[clean] = 0
            result.append(clean)
    return result, renamed


def _drop_garbage_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    # Vectorised — no Python loop over rows; safe for 100K-row chunks.
    str_df = df.fillna("").astype(str).apply(lambda s: s.str.strip())

    # 1. All-empty rows
    all_empty = (str_df == "").all(axis=1)

    # 2. First cell triggers a known garbage / subtotal keyword
    is_garbage = str_df.iloc[:, 0].str.match(_GARBAGE_ROW_RE, na=False)

    # 3. Separator rows: every non-empty cell consists entirely of separator chars.
    #    Logic: for each cell, (it is empty) OR (it matches the sep pattern).
    #    A row qualifies as separator only if it has at least one non-empty cell.
    non_empty_mask = str_df != ""
    sep_cell       = str_df.apply(lambda col: col.str.match(_SEP_ROW_RE, na=False))
    is_sep         = (~non_empty_mask | sep_cell).all(axis=1) & non_empty_mask.any(axis=1)

    mask   = ~(all_empty | is_garbage | is_sep)
    n_drop = int((~mask).sum())
    return df[mask].copy(), n_drop


# ══════════════════════════════════════════════════════════════════════════════
# Type detection (delegated to the detector registry)
# ══════════════════════════════════════════════════════════════════════════════
#
# All per-column type detection lives in app.services.preprocessor.type_detection.
# That module owns the IdentifierDetector, BooleanDetector, DateDetector and
# NumericDetector — each isolated, ordered, and unit-testable. Adding support
# for a new column type means adding one detector class, not editing this file.
#
# This wrapper exists only to (a) preserve the historical (col -> converter_fn)
# return shape consumed by the chunk-conversion code below, and (b) attach a
# human-readable warning trail for the per-file ingest log.

from app.services.preprocessor import detect_column_converter  # noqa: E402

ConverterFn = Callable[[object], object]


def _build_converters(
    sample: pd.DataFrame,
    headers: list[str],
    warns: list[str],
) -> dict[str, ConverterFn]:
    """Return the per-column converter map by consulting the detector registry.

    The registry's ordering rule is "identifier first, value-based last", so
    columns named like IDs (LEDGER_ID, INVOICE_NUM, ...) are never coerced —
    even when their values happen to look like years or numbers. Columns that
    no detector claims are intentionally left without a converter; the caller
    keeps them as raw strings.
    """
    converters: dict[str, ConverterFn] = {}
    for col in headers:
        if col not in sample.columns:
            continue
        sample_values = sample[col].dropna()
        result = detect_column_converter(col, sample_values)
        if result is None:
            continue
        if result.type_name == "identifier":
            warns.append(
                f"Column '{col}': preserved as identifier (no type coercion)"
            )
        converters[col] = result.convert
    return converters
